import pandas as pd
import re
from io import BytesIO
from google.colab import files
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import streamlit as st

# Load the product and rules data from uploaded Excel file
uploaded = st.file_uploader("Upload Product Detail Excel File", type="xlsx")
if uploaded:
    product_df = pd.read_excel(uploaded, sheet_name="Product detail")
    rules_df = pd.read_excel(uploaded, sheet_name="Rules")

    # Normalize column names
    product_df.columns = product_df.columns.str.strip()
    rules_df.columns = rules_df.columns.str.strip()

    # Sanity checks
    assert "TITLE" in product_df.columns, "Product file must contain a 'TITLE' column."
    assert all(col in rules_df.columns for col in ['Rule', 'Include', 'Exclude']), "Rules file must contain 'Rule', 'Include', and 'Exclude' columns."

    # Parse keywords with OR or AND logic
    def parse_keywords(text, is_exclude=False):
        if pd.isna(text) or not isinstance(text, str):
            return [], None
        text = text.strip().lower()
        logic = "or" if is_exclude else ("or" if " or " in text else "and")
        keywords = [w.strip() for w in re.split(r" and | or ", text) if w.strip()]
        return keywords, logic

    def preprocess_rules(rules_df):
        rule_dict = {}
        for _, row in rules_df.iterrows():
            rule = row['Rule']
            include_keywords, include_logic = parse_keywords(row['Include'])
            exclude_keywords, _ = parse_keywords(row['Exclude'], is_exclude=True)

            if rule not in rule_dict:
                rule_dict[rule] = {'include_blocks': [], 'exclude_keywords': set()}
            rule_dict[rule]['include_blocks'].append((include_keywords, include_logic))
            rule_dict[rule]['exclude_keywords'].update(exclude_keywords)
        return rule_dict

    def title_matches(title, include_blocks, exclude_keywords):
        if not isinstance(title, str):
            return False, []
        title_lower = title.lower()
        for keywords, logic in include_blocks:
            if logic == "and":
                if all(k in title_lower for k in keywords):
                    break
            elif logic == "or":
                if any(k in title_lower for k in keywords):
                    break
        else:
            return False, []  # No include block matched

        # Exclude check (OR logic)
        if any(k in title_lower for k in exclude_keywords):
            return False, []
        
        # Return matched keywords (for highlighting)
        matched_keywords = set()
        for keywords, _ in include_blocks:
            matched_keywords.update([k for k in keywords if k in title_lower])
        return True, list(matched_keywords)

    # Highlight the matched keywords with color
    highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    def highlight_keywords_with_color(ws, title_cell, keywords):
        if not isinstance(title_cell.value, str):
            return
        title = title_cell.value
        title_lower = title.lower()
        col = title_cell.column

        matched_so_far = 0  # Keep track of character matched so far to correctly slice original title
        for keyword in keywords:
            start_pos = title_lower.find(keyword.lower())
            while start_pos != -1:
                end_pos = start_pos + len(keyword)
                ws.cell(row=title_cell.row, column=col + 1, value=title[start_pos:end_pos]).fill = highlight_fill
                matched_so_far += len(keyword)
                start_pos = title_lower.find(keyword.lower(), end_pos)

    # Function to classify and expand
    def classify_and_expand(product_df, rules):
        results = []
        for _, row in product_df.iterrows():
            title = row['TITLE']
            matched = []
            matched_keywords_total = []
            for rule_name, rule_data in rules.items():
                matched_flag, matched_keywords = title_matches(title, rule_data['include_blocks'], rule_data['exclude_keywords'])
                if matched_flag:
                    matched.append(rule_name)
                    matched_keywords_total.extend(matched_keywords)
            if len(matched) == 0:
                results.append({**row, 'Rule match 1': '', 'Result': 'Exclude', 'TITLE (highlighted)': title})
            elif len(matched) == 1:
                results.append({**row, 'Rule match 1': matched[0], 'Result': 'Match', 'TITLE (highlighted)': title})
            else:
                for i, m in enumerate(matched):
                    results.append({**row, f'Rule match {i+1}': m, 'Result': 'Conflict' if i == 0 else '', 'TITLE (highlighted)': title})
        return pd.DataFrame(results)

    # Preprocess rules and classify
    parsed_rules = preprocess_rules(rules_df)
    output_df = classify_and_expand(product_df, parsed_rules)

    # Save the output to a workbook with formatting
    output_filename = "classified_products.xlsx"
    wb = Workbook()
    ws = wb.active

    # Write headers
    for col_num, column in enumerate(output_df.columns, 1):
        col_letter = get_column_letter(col_num)
        ws[f"{col_letter}1"] = column
        ws[f"{col_letter}1"].font = Font(bold=True)

    # Write data rows
    for row_num, row in output_df.iterrows():
        for col_num, value in enumerate(row, 1):
            cell = ws.cell(row=row_num + 2, column=col_num, value=value)

            # Apply color fill for 'TITLE (highlighted)' column
            if col_num == output_df.columns.get_loc('TITLE (highlighted)') + 1:
                if isinstance(cell.value, str):  # Simplified check
                    matched_keywords_for_highlighting = []
                    for rule_name, rule_data in parsed_rules.items():  # Find matched keywords again
                        matched_flag, matched_keywords = title_matches(cell.value, rule_data['include_blocks'], rule_data['exclude_keywords'])
                        if matched_flag:
                            matched_keywords_for_highlighting.extend(matched_keywords)
                    highlight_keywords_with_color(ws, cell, matched_keywords_for_highlighting)

    # Save the workbook
    wb.save(output_filename)

    # Allow the user to download the result
    st.download_button(label="Download Classified Products", data=open(output_filename, "rb"), file_name=output_filename, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

